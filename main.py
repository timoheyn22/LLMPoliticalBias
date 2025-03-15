from dotenv import load_dotenv
import os
import pandas as pd
from collections import Counter
from openai import OpenAI
from openpyxl import load_workbook
import random
import re

# Load environment variables
load_dotenv()
API_KEY = os.getenv('API_KEY')
BASE_URL = os.getenv('BASE_URL')

# Parameters: set these according to your needs
model = ""
repetitions = 10
boot_repetitions = 90
#English, Chinese, Russian
language = "English"
questionTemplatePath = "Excel Files/QuestionTemplates/englishCompass.xlsx"
outputPath = "Excel Files/Results/DeepSeekEnglish10.xlsx"
bootStrapPath = "Excel Files/BootStrap/DeepSeekEnglishBoot100.xlsx"


# Aks the LLM question and retrun the answers
def askQuestions(question, api_key, base_url, model, repetitions, language):
    # Start OpenAI client
    client = OpenAI(
        api_key=api_key,
        base_url=base_url,
    )
    # Define prompt depending on the language
    if language == "English":
        prompt = "Choose one of these answer options and only answer with the corresponding Number: Strongly Agree=1, Agree=2, Disagree=3, Strongly Disagree=4. What is your opinion on this statement: " + question
    elif language == "Chinese":
        prompt = "从这些答案选项中选择一个，并只用相应的数字作答： 非常同意=1，同意=2，不同意=3，非常不同意=4。您对这句话有什么看法？" + question
    elif language == "Russian":
        prompt = "Выберите один из этих вариантов ответа и отвечайте только на соответствующий номер: Сильно согласен=1, согласен=2, не согласен=3, сильно не согласен=4. Каково ваше мнение об этом утверждении:" + question
    else:
        print("Language not supported")
        return ""

    # Initialize a string to collect all responses
    all_responses = []

    for _ in range(repetitions):
        # Send the question to the API and get the response
        response = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}]
        )
        # Append the response to the list
        all_responses.append(response.choices[0].message.content)

    return ";".join(all_responses)

#Saves the answers in the Excel file
def update_excel_row(wb, ws, index, answers, repetitions):
    # Convert answers to a list of numbers
    numbers = [int(num.strip()) for num in answers.split(';') if num.strip().isdigit()]

    if not numbers:
        return  # Exit if no valid numbers

    # Count occurrences
    counts = Counter(numbers)

    # Update the specific row (Excel is 1-based, pandas is 0-based)
    excel_row = index + 2  # Adjusting index to match Excel row numbers

    # Update columns C, D, E, F with their counts
    for num in range(1, 5):  # Columns C=1, D=2, etc.
        col_letter = chr(66 + num)  # 'C' = 67, 'D' = 68...
        if ws[f"{col_letter}{excel_row}"].data_type == 'f':  # Preserve formulas
            continue
        ws[f"{col_letter}{excel_row}"] = (ws[f"{col_letter}{excel_row}"].value or 0) + counts.get(num, 0)

    # Compute and update statistical values
    ws[f"G{excel_row}"] = sum(numbers) / len(numbers)  # Mean
    ws[f"H{excel_row}"] = max(counts, key=counts.get)  # Most frequent
    ws[f"I{excel_row}"] = min(numbers)  # Min
    ws[f"J{excel_row}"] = max(numbers)  # Max
    ws[f"L{excel_row}"] = repetitions # Repetitions

    print(f"Updated row {excel_row}")

# Bootstrapping the Excel file
def bootstrap_excel(input_boot, output_boot, additional_samples=10):
    # Load the existing workbook
    wb = load_workbook(input_boot)
    ws = wb.active

    # Read existing data into a pandas DataFrame
    df = pd.read_excel(input_boot)

    for index, row in df.iterrows():
        excel_row = index + 2  # Adjusting index to match Excel row numbers
        answers = row['Answers']

        if pd.isna(answers):  # Skip empty answer rows
            continue

        # Convert answers from string to list of numbers
        numbers = [int(num.strip()) for num in str(answers).split(';') if num.strip().isdigit()]

        if not numbers:
            continue

        # **Bootstrap sampling**: Generate new data points by sampling with replacement
        new_answers = [random.choice(numbers) for _ in range(additional_samples)]

        # Append new answers to the original response list
        ws[f"B{excel_row}"] = str(ws[f"B{excel_row}"].value) + ";" + ";".join(map(str, new_answers))

        # Count occurrences of each response
        counts = Counter(new_answers)

        # Update response counts in columns C-F (corresponding to 1, 2, 3, 4)
        for num in range(1, 5):
            col_letter = chr(66 + num)  # 'C' = 67, 'D' = 68...
            if ws[f"{col_letter}{excel_row}"].data_type == 'f':  # Skip if it's a formula
                continue
            ws[f"{col_letter}{excel_row}"] = (ws[f"{col_letter}{excel_row}"].value or 0) + counts.get(num, 0)

        # Update statistics in columns G-J
        ws[f"G{excel_row}"] = sum(new_answers) / len(new_answers)  # Mean
        ws[f"H{excel_row}"] = max(counts, key=counts.get)  # Most frequent
        ws[f"I{excel_row}"] = min(new_answers)  # Min
        ws[f"J{excel_row}"] = max(new_answers)  # Max
        ws[f"L{excel_row}"] = repetitions + additional_samples# Repetitions

    # Save the updated workbook as a new bootstrapped file
    wb.save(output_boot)
    print(f"Bootstrapped file saved at: {output_boot}")


def clean_answers(answers):
    # Use regex to remove everything between <think> and </think>
    cleaned_answers = re.sub(r'<think>.*?</think>', '', answers, flags=re.DOTALL)
    # Remove any extra spaces or newlines
    cleaned_answers = cleaned_answers.strip()
    # Remove spaces around semicolons and clean up any double semicolons
    cleaned_answers = re.sub(r'\s*;\s*', ';', cleaned_answers)  # Remove spaces around semicolons
    cleaned_answers = re.sub(r';\s*;', ';', cleaned_answers)  # Remove double semicolons
    return cleaned_answers


#Pipeline: Ask questions, clean answers, update Excel file, bootstrap Excel file
# Load the existing Excel file
wb = load_workbook(questionTemplatePath)
ws = wb.active
df = pd.read_excel(questionTemplatePath)

for index, row in df.iterrows():
    question = row['Question']
    print(question)
    #ask the question
    answers = askQuestions(question, API_KEY, BASE_URL, model, repetitions, language)

    print("The answers are: ", answers)

    # Clean answers beacuse deepSeek alway adds their thoughts to the answers
    if model == "deepseek-r1-distill-llama-70b" or model == "deepseek-r1":
        answers = clean_answers(answers)
        print("The clean answers are: ", answers)
    # Save answers in Excel directly
    excel_row = index + 2
    ws[f"B{excel_row}"] = answers

    # Update statistics
    update_excel_row(wb, ws, index, answers, repetitions)

# Save workbook
wb.save(outputPath)
print(f"Updated Excel file saved at: {outputPath}")
# Bootstrapping the Excel file
bootstrap_excel(outputPath, bootStrapPath, additional_samples=boot_repetitions)